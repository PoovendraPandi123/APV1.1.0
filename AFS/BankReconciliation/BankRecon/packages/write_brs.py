import openpyxl
import shutil

excel_columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]

def template_fill(excel_file, sheet_name, cell_name, text):
    try:
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook[sheet_name]

        worksheet[cell_name] = text

        # worksheet_chg = worksheet[cell_name]
        # worksheet_chg.font = Font(color='00FF0000', bold=True)

        workbook.save(excel_file)
        workbook.close()
    except Exception as e:
        return {"Status": "Error", "Message": str(e)}

def template_cell_fill(page):
    try:
        workbook = openpyxl.load_workbook(page["excel_file"])
        worksheet = workbook[page["sheet_name"]]

        for data in page["data"]:
            # print(data["cell_number"])
            # print(data["cell_value"])
            worksheet[data["cell_number"]] = data["cell_value"]

        workbook.save(page["excel_file"])
        workbook.close()
    except Exception as e:
        print("Error")
        print(e)
        return {"Status": "Error", "Message": str(e)}

def template_multiple_fill(excel_file, sheet_name, cell_name, cell_start_number, data_list):
    try:
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook[sheet_name]

        # cell_position = cell_name + str(cell_start_number)

        for i in range(0, len(data_list)):
            # print(cell_name + str(cell_start_number))
            worksheet[cell_name + str(cell_start_number)] = data_list[i]
            cell_start_number = cell_start_number + 1

        workbook.save(excel_file)
        workbook.close()
    except Exception as e:
        print("Error")
        print(e)
        print(data_list[i])
        print(cell_start_number)
        print(cell_name)
        print({"Status": "Error", "Message": str(e)})
        return {"Status": "Error", "Message": str(e)}

def multiple_cell_fill():
    try:
        pass
    except Exception as e:
        return {"Status": "Error", "Message": str(e)}

def write_brs_file(data):
    try:
        excel_file = "G:/AdventsProduct/V1.0.0/AFS/Reconciliation/static/BRS-Report-Template.xlsx"

        new_file = "G:/AdventsProduct/V1.0.0/AFS/Reconciliation/static/" + "Report-" + data["relationship"] + "-" + data["report_generation_count"] + ".xlsx"

        shutil.copy(excel_file, new_file)


        # For Home Page
        home_page = {
            "excel_file" : new_file,
            "sheet_name": "Home",
            "data" : [
                {"cell_number": "D5", "cell_value": data["company_name"]},
                {"cell_number": "D6", "cell_value": data["report_generation_date"]},
                {"cell_number": "D9", "cell_value": data["relationship"]},
                {"cell_number": "D10", "cell_value": data["report_date"]},
                {"cell_number": "D12", "cell_value": data["bank_name"]},
                {"cell_number": "D13", "cell_value": data["branch_location"]},
                {"cell_number": "D14", "cell_value": data["account_type"]},
                {"cell_number": "D15", "cell_value": data["account_number"]},
            ]
        }

        template_cell_fill(home_page)

        # template_fill(excel_file, "Home", "D5", data["company_name"])
        # template_fill(excel_file, "Home", "D6", data["report_generation_date"])
        # template_fill(excel_file, "Home", "D12", data["bank_name"])
        # template_fill(excel_file, "Home", "D13", data["branch_location"])
        # template_fill(excel_file, "Home", "D14", data["account_type"])
        # template_fill(excel_file, "Home", "D15", data["account_number"])

        # For Details
        details_page = {
            "excel_file": new_file,
            "sheet_name": "Details",
            "data" : [
                {"cell_number": "B5", "cell_value": data["accpac_code"]},
                {"cell_number": "B6", "cell_value": data["bank_code"]},
                {"cell_number": "B7", "cell_value": data["report_generation_date"]},
                {"cell_number": "B8", "cell_value": data["report_date"]},
                {"cell_number": "B9", "cell_value": data["bank_closing_balance"]},
                {"cell_number": "B10", "cell_value": data["gl_closing_balance"]},
            ]
        }

        template_cell_fill(details_page)

        # For GL Debits and Credits
        for i in range(0, len(data["gl_debits_output"].columns)):
            template_multiple_fill(new_file, "Cheques & Letters OS", excel_columns[i], 8, data["gl_debits_output"][i])

        # For Bank Debits and Credits
        for i in range(0, len(data["bank_debits_output"].columns)):
            template_multiple_fill(new_file, "Debited & Credited by Bank", excel_columns[i], 8, data["bank_debits_output"][i])

        # For GL Records
        gl_records_page = {
            "excel_file": new_file,
            "sheet_name": "GL Records",
            "data" : [
                {"cell_number": "W2", "cell_value": data["rep_gl_opening_balance"]},
            ]
        }
        template_cell_fill(gl_records_page)

        for i in range(0, len(data["rep_gl_table_query_output"].columns)):
            template_multiple_fill(new_file, "GL Records", excel_columns[i], 3, data["rep_gl_table_query_output"][i])

        # For Bank Records
        bank_records_page = {
            "excel_file": new_file,
            "sheet_name": "Bank Statement",
            "data" : [
                {"cell_number": "H4", "cell_value": data["rep_bank_opening_balance"]},
            ]
        }
        template_cell_fill(bank_records_page)

        for i in range(0, len(data["rep_bank_table_query_output"].columns)):
            template_multiple_fill(new_file, "Bank Statement", excel_columns[i], 5, data["rep_bank_table_query_output"][i])


        # print(data["gl_debits_output"][0])
        # print(len(data["gl_debits_output"]))

        # cell_value = 8
        # for i in range(0, len(data["gl_debits_output"])):
        #     template_fill(excel_file, "Cheques & Letters OS", "A"+str(cell_value), data["gl_debits_output"][0][i])
        #     cell_value += 1
        # workbook = openpyxl.load_workbook('C:/AdventsProduct/V1.0.0/AFS/Reports/TeamLease-BRS-Report-Staffing.xlsx')
        # writer = pd.ExcelWriter('C:/AdventsProduct/V1.0.0/AFS/Reports/TeamLease-BRS-Report-Staffing.xlsx', engine="openpyxl")
        # writer.book = workbook
        # data["gl_debits_output"].to_excel(writer, sheet_name="Cheques & Letters OS", index=False, startrow=9, startcol=1, engine="openpyxl")
        # writer.save()
        # writer.close()

        file_generated = "http://localhost:50004/static/"+new_file.split("/")[-1]

        return {"Status": "Success", "file_generated": file_generated}
    except Exception as e:
        # print(e)
        return {"Status": "Error", "Message": str(e)}

def write_consolidation_file(data):
    try:
        excel_file = "G:/AdventsProduct/V1.0.0/AFS/Reconciliation/static/ConsolidationReport/Consolidation-Report-Template.xlsx"

        new_file = "G:/AdventsProduct/V1.0.0/AFS/Reconciliation/static/ConsolidationReport/" + "Consolidation-Report-" + data["report_generation_count"] + ".xlsx"

        shutil.copy(excel_file, new_file)

        # For Consolidation
        # print(data["gl_closing_balance_hdfc062"])
        consolidation_page = {
            "excel_file": new_file,
            "sheet_name": "Consolidation",
            "data" : [
                {"cell_number": "B8", "cell_value": data["gl_closing_balance_hdfc062"]},
                {"cell_number": "B10", "cell_value": data["bank_closing_balance_hdfc062"]},
                {"cell_number": "B16", "cell_value": data["gl_credit_hdfc062"]},
                {"cell_number": "B17", "cell_value": data["bank_credit_hdfc062"]},
                {"cell_number": "B19", "cell_value": data["bank_debit_hdfc062"]},
                {"cell_number": "B20", "cell_value": data["gl_debit_hdfc062"]},
                {"cell_number": "C8", "cell_value": data["gl_closing_balance_hdfc295"]},
                {"cell_number": "C10", "cell_value": data["bank_closing_balance_hdfc295"]},
                {"cell_number": "C16", "cell_value": data["gl_credit_hdfc295"]},
                {"cell_number": "C17", "cell_value": data["bank_credit_hdfc295"]},
                {"cell_number": "C19", "cell_value": data["bank_debit_hdfc295"]},
                {"cell_number": "C20", "cell_value": data["gl_debit_hdfc295"]},
                {"cell_number": "D8", "cell_value": data["gl_closing_balance_sbi90642"]},
                {"cell_number": "D10", "cell_value": data["bank_closing_balance_sbi90642"]},
                {"cell_number": "D16", "cell_value": data["gl_credit_sbi90642"]},
                {"cell_number": "D17", "cell_value": data["bank_credit_sbi90642"]},
                {"cell_number": "D19", "cell_value": data["bank_debit_sbi90642"]},
                {"cell_number": "D20", "cell_value": data["gl_debit_sbi90642"]},
                {"cell_number": "E8", "cell_value": data["gl_closing_balance_icici2605"]},
                {"cell_number": "E10", "cell_value": data["bank_closing_balance_icici2605"]},
                {"cell_number": "E16", "cell_value": data["gl_credit_icici2605"]},
                {"cell_number": "E17", "cell_value": data["bank_credit_icici2605"]},
                {"cell_number": "E19", "cell_value": data["bank_debit_icici2605"]},
                {"cell_number": "E20", "cell_value": data["gl_debit_icici2605"]},
                {"cell_number": "F8", "cell_value": data["gl_closing_balance_axis18294"]},
                {"cell_number": "F10", "cell_value": data["bank_closing_balance_axis18294"]},
                {"cell_number": "F16", "cell_value": data["gl_credit_axis18294"]},
                {"cell_number": "F17", "cell_value": data["bank_credit_axis18294"]},
                {"cell_number": "F19", "cell_value": data["bank_debit_axis18294"]},
                {"cell_number": "F20", "cell_value": data["gl_debit_axis18294"]},
                {"cell_number": "G8", "cell_value": data["gl_closing_balance_hdfc828"]},
                {"cell_number": "G10", "cell_value": data["bank_closing_balance_hdfc828"]},
                {"cell_number": "G16", "cell_value": data["gl_credit_hdfc828"]},
                {"cell_number": "G17", "cell_value": data["bank_credit_hdfc828"]},
                {"cell_number": "G19", "cell_value": data["bank_debit_hdfc828"]},
                {"cell_number": "G20", "cell_value": data["gl_debit_hdfc828"]},
                {"cell_number": "H8", "cell_value": data["gl_closing_balance_idbi1946"]},
                {"cell_number": "H10", "cell_value": data["bank_closing_balance_idbi1946"]},
                {"cell_number": "H16", "cell_value": data["gl_credit_idbi1946"]},
                {"cell_number": "H17", "cell_value": data["bank_credit_idbi1946"]},
                {"cell_number": "H19", "cell_value": data["bank_debit_idbi1946"]},
                {"cell_number": "H20", "cell_value": data["gl_debit_idbi1946"]},
                {"cell_number": "I8", "cell_value": data["gl_closing_balance_kotak09195"]},
                {"cell_number": "I10", "cell_value": data["bank_closing_balance_kotak09195"]},
                {"cell_number": "I16", "cell_value": data["gl_credit_kotak09195"]},
                {"cell_number": "I17", "cell_value": data["bank_credit_kotak09195"]},
                {"cell_number": "I19", "cell_value": data["bank_debit_kotak09195"]},
                {"cell_number": "I20", "cell_value": data["gl_debit_kotak09195"]},
                {"cell_number": "J8", "cell_value": data["gl_closing_balance_fedral6006"]},
                {"cell_number": "J10", "cell_value": data["bank_closing_balance_fedral6006"]},
                {"cell_number": "J16", "cell_value": data["gl_credit_fedral6006"]},
                {"cell_number": "J17", "cell_value": data["bank_credit_fedral6006"]},
                {"cell_number": "J19", "cell_value": data["bank_debit_fedral6006"]},
                {"cell_number": "J20", "cell_value": data["gl_debit_fedral6006"]},
                {"cell_number": "K8", "cell_value": data["gl_closing_balance_axis1710"]},
                {"cell_number": "K10", "cell_value": data["bank_closing_balance_axis1710"]},
                {"cell_number": "K16", "cell_value": data["gl_credit_axis1710"]},
                {"cell_number": "K17", "cell_value": data["bank_credit_axis1710"]},
                {"cell_number": "K19", "cell_value": data["bank_debit_axis1710"]},
                {"cell_number": "K20", "cell_value": data["gl_debit_axis1710"]},
                {"cell_number": "L8", "cell_value": data["gl_closing_balance_kotak12521"]},
                {"cell_number": "L10", "cell_value": data["bank_closing_balance_kotak12521"]},
                {"cell_number": "L16", "cell_value": data["gl_credit_kotak12521"]},
                {"cell_number": "L17", "cell_value": data["bank_credit_kotak12521"]},
                {"cell_number": "L19", "cell_value": data["bank_debit_kotak12521"]},
                {"cell_number": "L20", "cell_value": data["gl_debit_kotak12521"]},
                {"cell_number": "M8", "cell_value": data["gl_closing_balance_indus33974"]},
                {"cell_number": "M10", "cell_value": data["bank_closing_balance_indus339740"]},
                {"cell_number": "M16", "cell_value": data["gl_credit_indus33974"]},
                {"cell_number": "M17", "cell_value": data["bank_credit_indus33974"]},
                {"cell_number": "M19", "cell_value": data["bank_debit_indus33974"]},
                {"cell_number": "M20", "cell_value": data["gl_debit_indus33974"]},
                {"cell_number": "N8", "cell_value": data["gl_closing_balance_fedral5979"]},
                {"cell_number": "N10", "cell_value": data["bank_closing_balance_fedral5979"]},
                {"cell_number": "N16", "cell_value": data["gl_credit_fedral5979"]},
                {"cell_number": "N17", "cell_value": data["bank_credit_fedral5979"]},
                {"cell_number": "N19", "cell_value": data["bank_debit_fedral5979"]},
                {"cell_number": "N20", "cell_value": data["gl_debit_fedral5979"]},
                {"cell_number": "O8", "cell_value": data["gl_closing_balance_hdfccms828"]},
                {"cell_number": "O10", "cell_value": data["bank_closing_balance_hdfccms828"]},
                {"cell_number": "O16", "cell_value": data["gl_credit_hdfccms828"]},
                {"cell_number": "O17", "cell_value": data["bank_credit_hdfccms828"]},
                {"cell_number": "O19", "cell_value": data["bank_debit_hdfccms828"]},
                {"cell_number": "O20", "cell_value": data["gl_debit_hdfccms828"]},
                {"cell_number": "P8", "cell_value": data["gl_closing_balance_kotak18019"]},
                {"cell_number": "P10", "cell_value": data["bank_closing_balance_kotak18019"]},
                {"cell_number": "P16", "cell_value": data["gl_credit_kotak18019"]},
                {"cell_number": "P17", "cell_value": data["bank_credit_kotak18019"]},
                {"cell_number": "P19", "cell_value": data["bank_debit_kotak18019"]},
                {"cell_number": "P20", "cell_value": data["gl_debit_kotak18019"]},
                {"cell_number": "Q8", "cell_value": data["gl_closing_balance_axis35274"]},
                {"cell_number": "Q10", "cell_value": data["bank_closing_balance_axis35274"]},
                {"cell_number": "Q16", "cell_value": data["gl_credit_axis35274"]},
                {"cell_number": "Q17", "cell_value": data["bank_credit_axis35274"]},
                {"cell_number": "Q19", "cell_value": data["bank_debit_axis35274"]},
                {"cell_number": "Q20", "cell_value": data["gl_debit_axis35274"]},
                {"cell_number": "R8", "cell_value": data["gl_closing_balance_icici240"]},
                {"cell_number": "R10", "cell_value": data["bank_closing_balance_icici240"]},
                {"cell_number": "R16", "cell_value": data["gl_credit_icici240"]},
                {"cell_number": "R17", "cell_value": data["bank_credit_icici240"]},
                {"cell_number": "R19", "cell_value": data["bank_debit_icici240"]},
                {"cell_number": "R20", "cell_value": data["gl_debit_icici240"]},
                {"cell_number": "S8", "cell_value": data["gl_closing_balance_rbl0110"]},
                {"cell_number": "S10", "cell_value": data["bank_closing_balance_rbl0110"]},
                {"cell_number": "S16", "cell_value": data["gl_credit_rbl0110"]},
                {"cell_number": "S17", "cell_value": data["bank_credit_rbl0110"]},
                {"cell_number": "S19", "cell_value": data["bank_debit_rbl0110"]},
                {"cell_number": "S20", "cell_value": data["gl_debit_rbl0110"]},
            ]
        }

        template_cell_fill(consolidation_page)

        file_generated = "http://localhost:50004/static/ConsolidationReport/" + new_file.split("/")[-1]

        return {"Status": "Success", "file_generated": file_generated}
    except Exception as e:
        return {"Status": "Error", "Message": str(e)}